"""Shared best-effort text extraction for Office/PDF binary formats.

Owned by: ingestion/. Originally written inside `connectors/sharepoint.py`
(Milestone 9's SharePoint connector, closing that connector's own
"plain-text only" gap); pulled out here once `connectors/confluence.py`
needed the exact same capability for attachment content -- a shared,
source-agnostic parsing utility, not per-connector business logic, so it
lives as a peer to `connectors/`/`processors/`/`schemas.py`, not inside
either connector.

One public entry point, `extract_text`, dispatches by file extension and
returns `None` for an unsupported extension or a parse failure (corrupt,
password-protected, actually a different format than its extension claims,
...) -- the same "skipped, not an error" contract every connector's own
undecodable-content handling already follows (e.g.
`GitHubConnector._fetch_file_content`'s undecodable-file treatment). A
single bad file must never fail an entire ingestion job.

Each extractor is a best-effort text join with no layout/formatting
preserved (PDF: every page's `extract_text()`; DOCX: every paragraph;
XLSX: every non-empty cell, read-only mode so a large workbook streams
rather than loading fully into memory) -- callers want ingestible search
content, not a faithful document rendering.
"""

from __future__ import annotations

from io import BytesIO

import docx
from openpyxl import load_workbook
from pypdf import PdfReader

from app.shared.config.logging import get_logger

logger = get_logger(__name__)


def _extract_pdf_text(raw_bytes: bytes) -> str:
    reader = PdfReader(BytesIO(raw_bytes))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx_text(raw_bytes: bytes) -> str:
    document = docx.Document(BytesIO(raw_bytes))
    return "\n\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text)


def _extract_xlsx_text(raw_bytes: bytes) -> str:
    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        lines: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(value) for value in row if value is not None]
                if cells:
                    lines.append("\t".join(cells))
        return "\n".join(lines)
    finally:
        workbook.close()


_EXTRACTORS = {
    ".pdf": _extract_pdf_text,
    ".docx": _extract_docx_text,
    ".xlsx": _extract_xlsx_text,
}


def extract_text(filename: str, raw_bytes: bytes) -> str | None:
    """Extract text from `raw_bytes`, dispatched by `filename`'s extension.

    Returns `None` if `filename`'s extension isn't one of `_EXTRACTORS`, or
    if parsing raises for any reason -- callers should treat `None` as
    "skip this item," never as a reason to fail the whole batch.
    """
    extension = next((ext for ext in _EXTRACTORS if filename.endswith(ext)), None)
    if extension is None:
        return None
    try:
        return _EXTRACTORS[extension](raw_bytes)
    except Exception:
        logger.warning("office_extraction_failed", filename=filename, extension=extension)
        return None
